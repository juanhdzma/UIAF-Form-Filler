from openpyxl import load_workbook

# Constants
NATURAL_TYPES = ["C", "E", "P", "T", "NAN"]
JURIDICAL_TYPES = ["N", "NAJ"]

LEGAL_TYPE_MAP = {"N": "NIT", "NAJ": "Sociedad extranjera sin NIT en Colombia"}

NATURAL_TYPE_MAP = {
    "C": "Cédula de ciudadanía",
    "E": "Cédula de extranjería",
    "P": "Pasaporte",
    "T": "Tarjeta de identidad",
    "NAN": "Documento de identificación extranjero",
}


def load_templates(tipo):
    if tipo == "NATURAL":
        return load_workbook("src/templates/Naturales.xlsx")
    elif tipo == "JURIDICA":
        return load_workbook("src/templates/Juridicas.xlsx")


def create_excel_file(rows, tipo):
    excel = load_templates(tipo)
    sheet = excel.active

    for i, row in enumerate(rows, start=4):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)

    excel.save(f"output_files/{tipo}.xlsx")
    excel.close()


def process_legal_entities(df):
    document_counter = 0
    rl_counter = 0
    rows = []

    for i, row in df.iterrows():
        tipo = LEGAL_TYPE_MAP.get(str(row["Tipo de documento"]).upper(), "")
        documento = str(row["Documento "])

        if documento.strip() == "":
            documento = document_counter
            document_counter += 1

        rl_counter += 1
        nombre = row["Nombre"]
        relacion = row["Relación"]

        row_data = [
            tipo,
            documento,
            nombre,
            "NA",
            "NA",
            "NA",
            "Cédula de ciudadanía",
            f"{i+1}",
            relacion,
            "NO",
        ]
        rows.append(row_data)

    return rows


def process_natural_entities(df):
    document_counter = 0
    rows = []

    for _, row in df.iterrows():
        tipo_raw = str(row["Tipo de documento"]).upper()
        tipo = NATURAL_TYPE_MAP.get(tipo_raw, "DESCONOCIDO")

        documento = str(row["Documento "])
        if documento.strip() == "":
            documento = document_counter
            document_counter += 1

        nombre = row["Nombre"]
        relacion = row["Relación"]

        partes = str(nombre).split()

        if len(partes) >= 4:
            primer_nombre = " ".join(partes[:-3])
            primer_apellido = partes[-2]
            segundo_apellido = partes[-1]
        elif len(partes) == 3:
            primer_nombre = partes[0]
            primer_apellido = partes[1]
            segundo_apellido = partes[2]
        elif len(partes) == 2:
            primer_nombre = partes[0]
            primer_apellido = partes[1]
            segundo_apellido = ""
        else:
            primer_nombre = nombre
            primer_apellido = ""
            segundo_apellido = ""

        row_data = [
            tipo,
            documento,
            primer_nombre,
            primer_apellido,
            segundo_apellido,
            relacion,
            "NO",
        ]
        rows.append(row_data)

    return rows


def process_relation_df(df, tipo):
    if tipo == "NATURAL":
        valores = NATURAL_TYPES
        df_filtrado = df[df["Tipo de documento"].isin(valores)]
        rows = process_natural_entities(df_filtrado)
    elif tipo == "JURIDICA":
        valores = JURIDICAL_TYPES
        df_filtrado = df[df["Tipo de documento"].isin(valores)]
        rows = process_legal_entities(df_filtrado)

    create_excel_file(rows, tipo)
